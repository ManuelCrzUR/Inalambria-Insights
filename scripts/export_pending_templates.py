#!/usr/bin/env python3
"""
export_pending_templates.py - Exporta top 100 plantillas pending (no clasificadas)

Genera XLSX con plantillas que no fueron clasificadas por reglas L0,
para análisis manual y creación de nuevas reglas.

Uso:
    python3 scripts/export_pending_templates.py 2026-05-10
    python3 scripts/export_pending_templates.py 2026-05-10 --top 50
    python3 scripts/export_pending_templates.py 2026-05-10 --top 200
"""

import sys
import json
import argparse
from pathlib import Path
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Error: openpyxl no está instalado")
    print("Instálalo con: pip install openpyxl")
    sys.exit(1)


def export_pending_templates(output_date: str, top_n: int = 100, output_dir: Path = None):
    """Exporta las top N plantillas pending a XLSX."""

    if output_dir is None:
        output_dir = Path("/home/manuel-cruz/Desktop/Twnel/prod_pipeline/output") / output_date

    rule_classifications_path = output_dir / "rule_classifications.jsonl"

    if not rule_classifications_path.exists():
        print(f"❌ Archivo no encontrado: {rule_classifications_path}")
        return None

    pending_templates = []
    total_lines = 0
    pending_count = 0

    print(f"📖 Leyendo {rule_classifications_path}...")

    with open(rule_classifications_path, 'r', encoding='utf-8') as f:
        for line in f:
            total_lines += 1

            try:
                result = json.loads(line)

                # Solo plantillas pending (sin clasificar)
                if result.get('level_used') == 'pending':
                    pending_count += 1

                    pending_templates.append({
                        'template_id': result.get('template_id', ''),
                        'template_text': result.get('template_text', ''),
                        'frequency': result.get('frequency', 1),
                        'applied_rules': ', '.join(result.get('applied_rules', [])),
                    })

                    if pending_count % 500000 == 0:
                        print(f"  ✓ Procesadas {pending_count:,} plantillas pending...")

            except json.JSONDecodeError as e:
                print(f"⚠️  Error JSON en línea {total_lines}: {e}")
                continue

    print(f"\n✅ Análisis completado:")
    print(f"   Total de líneas: {total_lines:,}")
    print(f"   Plantillas pending: {pending_count:,}")

    # Ordenar por frecuencia descendente y tomar top N
    pending_sorted = sorted(pending_templates, key=lambda x: x['frequency'], reverse=True)
    top_pending = pending_sorted[:top_n]

    print(f"\n📊 Top {top_n} plantillas pending por frecuencia:")
    for idx, data in enumerate(top_pending[:10], 1):
        print(f"  {idx}. {data['template_text'][:60]:60s} | freq={data['frequency']:,}")

    return top_pending, output_dir


def generate_xlsx(pending_templates, top_n: int, output_dir: Path):
    """Genera XLSX con plantillas pending."""
    xlsx_path = output_dir / f"pending_templates_top{top_n}.xlsx"

    print(f"\n📊 Generando XLSX: {xlsx_path}")

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pending Templates"

    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(size=11)
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Encabezados
    headers = ["#", "ID Plantilla", "Ejemplo (Template Text)", "Frecuencia", "Reglas Aplicadas"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Datos
    for row_num, data in enumerate(pending_templates, 2):
        cells_data = [
            row_num - 1,  # Número de orden
            data['template_id'],
            data['template_text'],
            data['frequency'],
            data['applied_rules'],
        ]

        for col_num, value in enumerate(cells_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

            # Centrar columnas numéricas
            if col_num in [1, 4]:  # # y Frecuencia
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 5  # #
    ws.column_dimensions['B'].width = 20  # ID
    ws.column_dimensions['C'].width = 60  # Template Text
    ws.column_dimensions['D'].width = 15  # Frecuencia
    ws.column_dimensions['E'].width = 40  # Reglas

    # Altura de filas
    ws.row_dimensions[1].height = 30  # Header

    # Freeze panes (congelar encabezado)
    ws.freeze_panes = "A2"

    # Guardar
    wb.save(xlsx_path)
    print(f"✅ XLSX guardado en: {xlsx_path}")

    return xlsx_path


def generate_summary_sheet(pending_templates, output_dir: Path):
    """Genera XLSX con resumen adicional."""
    xlsx_path = output_dir / "pending_templates_analysis.xlsx"

    print(f"📊 Generando análisis en XLSX: {xlsx_path}")

    from collections import Counter

    wb = Workbook()

    # Sheet 1: Top templates
    ws_top = wb.active
    ws_top.title = "Top 100 Pending"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Encabezados
    headers = ["Rango", "ID Plantilla", "Ejemplo", "Frecuencia", "% del Total"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_top.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    total_freq = sum(t['frequency'] for t in pending_templates)

    # Datos
    for row_num, data in enumerate(pending_templates, 2):
        pct = (data['frequency'] / total_freq * 100) if total_freq > 0 else 0

        cells_data = [
            row_num - 1,
            data['template_id'],
            data['template_text'],
            data['frequency'],
            f"{pct:.2f}%",
        ]

        for col_num, value in enumerate(cells_data, 1):
            cell = ws_top.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [1, 4, 5]:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws_top.column_dimensions['A'].width = 8
    ws_top.column_dimensions['B'].width = 20
    ws_top.column_dimensions['C'].width = 60
    ws_top.column_dimensions['D'].width = 15
    ws_top.column_dimensions['E'].width = 12

    ws_top.freeze_panes = "A2"

    # Sheet 2: Estadísticas
    ws_stats = wb.create_sheet("Estadísticas")

    stats_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    stats_header_font = Font(bold=True, color="FFFFFF", size=11)

    # Título
    ws_stats['A1'] = "Estadísticas de Plantillas Pending"
    ws_stats['A1'].font = Font(bold=True, size=14)

    # Métricas
    metrics = [
        ("Total Plantillas Pending", len(pending_templates)),
        ("Frecuencia Total", total_freq),
        ("Frecuencia Promedio", total_freq / len(pending_templates) if pending_templates else 0),
        ("Frecuencia Máxima", max((t['frequency'] for t in pending_templates), default=0)),
        ("Frecuencia Mínima", min((t['frequency'] for t in pending_templates), default=0)),
    ]

    for row_num, (label, value) in enumerate(metrics, 3):
        ws_stats[f'A{row_num}'] = label
        ws_stats[f'A{row_num}'].font = Font(bold=True)
        ws_stats[f'B{row_num}'] = value if isinstance(value, str) else f"{value:,.0f}"
        ws_stats[f'B{row_num}'].font = Font(size=11)

    # Análisis por reglas aplicadas
    ws_stats['A9'] = "Reglas más frecuentes en plantillas pending:"
    ws_stats['A9'].font = Font(bold=True, size=12)

    rules_counter = Counter()
    for template in pending_templates:
        if template['applied_rules']:
            for rule in template['applied_rules'].split(', '):
                if rule.strip():
                    rules_counter[rule.strip()] += 1

    for idx, (rule, count) in enumerate(rules_counter.most_common(20), 11):
        ws_stats[f'A{idx}'] = rule
        ws_stats[f'B{idx}'] = count

    ws_stats.column_dimensions['A'].width = 35
    ws_stats.column_dimensions['B'].width = 20

    # Guardar
    wb.save(xlsx_path)
    print(f"✅ Análisis guardado en: {xlsx_path}")

    return xlsx_path


def main():
    parser = argparse.ArgumentParser(
        description="Exporta top N plantillas pending a XLSX"
    )
    parser.add_argument(
        'output_date',
        help='Fecha del output (ej: 2026-05-10)'
    )
    parser.add_argument(
        '--top',
        type=int,
        default=100,
        help='Número de top plantillas (default: 100)'
    )

    args = parser.parse_args()

    print(f"\n{'='*70}")
    print(f"📋 Exportando plantillas pending (no clasificadas)")
    print(f"{'='*70}\n")

    result = export_pending_templates(args.output_date, top_n=args.top)

    if result is None:
        return 1

    pending_templates, output_dir = result

    if not pending_templates:
        print("\n⚠️  No hay plantillas pending.")
        return 1

    # Generar XLSX
    generate_xlsx(pending_templates, args.top, output_dir)
    generate_summary_sheet(pending_templates, output_dir)

    print(f"\n✅ Exportación completada exitosamente!\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())
